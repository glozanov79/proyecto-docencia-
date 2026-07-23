"""
Genera el taller (.xlsx) a partir de un brief.md, llamando a la API de Claude
para diseñar los ejercicios.

Requiere:
- Variable de entorno ANTHROPIC_API_KEY

Uso:
    python generar_taller.py salidas/2026-08-05/brief.md
"""

import json
import os
import re
import sys
from pathlib import Path

import anthropic
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

MODEL = "claude-haiku-4-5-20251001"

SYSTEM_PROMPT = """Eres un diseñador de talleres académicos para un curso universitario.
Vas a recibir el brief de una clase y debes diseñar un taller de 8 a 10 ejercicios
para que los estudiantes practiquen en Moodle, coherente con la materia y el nivel
del curso descrito en el brief.

REGLAS:
- Los ejercicios deben cubrir TODOS los subtemas de "Contenido clave a cubrir" del brief,
  no solo el primero.
- Nivel introductorio/intermedio, coherente con una clase presencial de duración normal.
- Cada ejercicio debe ser autónomo (con los datos numéricos o el enunciado completo),
  nunca una instrucción genérica como "resuelve un ejercicio de X".
- Responde ÚNICAMENTE con JSON válido, sin texto adicional ni markdown, con esta forma
  exacta:
  {"titulo": "Taller - Semana N", "ejercicios": [{"numero": 1, "tema": "...", "enunciado": "..."}]}
"""


def extraer_json(texto):
    texto = texto.strip()
    if texto.startswith("```"):
        texto = re.sub(r"^```(json)?", "", texto).rstrip("`").strip()
    inicio = texto.find("{")
    fin = texto.rfind("}")
    return json.loads(texto[inicio:fin + 1])


def generar_taller(brief_texto, api_key):
    client = anthropic.Anthropic(api_key=api_key)
    respuesta = client.messages.create(
        model=MODEL,
        max_tokens=4000,
        system=SYSTEM_PROMPT,
        messages=[
            {"role": "user", "content": f"Brief de la clase:\n\n{brief_texto}\n\nGenera el taller en JSON."}
        ],
    )
    texto = "".join(b.text for b in respuesta.content if b.type == "text")
    return extraer_json(texto)


def escribir_xlsx(datos, ruta_salida):
    wb = Workbook()
    ws = wb.active
    ws.title = "Taller"

    ws["A1"] = datos.get("titulo", "Taller")
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:C1")

    for col, texto in enumerate(["#", "Tema", "Enunciado"], start=1):
        ws.cell(row=3, column=col, value=texto).font = Font(bold=True)

    for i, ej in enumerate(datos.get("ejercicios", []), start=4):
        ws.cell(row=i, column=1, value=ej.get("numero"))
        ws.cell(row=i, column=2, value=ej.get("tema", ""))
        celda = ws.cell(row=i, column=3, value=ej.get("enunciado", ""))
        celda.alignment = Alignment(wrap_text=True, vertical="top")

    for col, ancho in {1: 6, 2: 28, 3: 90}.items():
        ws.column_dimensions[get_column_letter(col)].width = ancho

    wb.save(ruta_salida)


def main():
    if len(sys.argv) != 2:
        print("Uso: python generar_taller.py <ruta_al_brief.md>")
        sys.exit(1)

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("ERROR: define la variable de entorno ANTHROPIC_API_KEY antes de correr esto.")
        sys.exit(1)

    brief_path = Path(sys.argv[1])
    brief_texto = brief_path.read_text(encoding="utf-8")
    salida_dir = brief_path.parent

    print("Llamando a la API de Claude para diseñar el taller...")
    datos = generar_taller(brief_texto, api_key)

    ruta_salida = salida_dir / "taller.xlsx"
    escribir_xlsx(datos, ruta_salida)
    print(f"taller.xlsx generado en: {ruta_salida}")


if __name__ == "__main__":
    main()
